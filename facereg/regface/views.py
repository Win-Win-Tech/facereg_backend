import base64
import logging
import os
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal

import face_recognition
import numpy as np
from django.conf import settings
from django.db.models import Min, Max
from django.utils import timezone
from django.utils.timezone import localtime
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None
from openpyxl import Workbook
from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView

from .face_utils import get_face_encoding
from .models import AttendanceLog, AuthToken, Employee, Location, PayrollRecord, User
from .serializers import (
    EmployeeRegisterSerializer,
    EmployeeListSerializer,
    EmployeeSerializer,
    EmployeeUpdateSerializer,
    FaceUploadSerializer,
    LocationSerializer,
    UserSerializer,
)
from .authentication import SimpleTokenAuthentication

logger = logging.getLogger(__name__)


def _ensure_aware(dt):
    if dt is None:
        return None
    if dt.tzinfo is None:
        return timezone.make_aware(dt)
    return dt


def to_ist(dt):
    """Convert a datetime to Asia/Kolkata timezone and return a datetime.

    Accepts naive or aware datetimes. Returns None if input is None.
    """
    if dt is None:
        return None
    dt = _ensure_aware(dt)
    if ZoneInfo is not None:
        return dt.astimezone(ZoneInfo("Asia/Kolkata"))
    # fallback to pytz if zoneinfo is not available
    try:
        import pytz

        tz = pytz.timezone("Asia/Kolkata")
        return dt.astimezone(tz)
    except Exception:
        return dt.astimezone(timezone.get_current_timezone())


def is_superadmin(user: User) -> bool:
    return getattr(user, "role", None) == User.Role.SUPERADMIN


class AuthenticatedAPIView(APIView):
    authentication_classes = [SimpleTokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]


class LocationListCreateView(AuthenticatedAPIView):
    def get(self, request):
        include_deleted = request.query_params.get("include_deleted") == "true"
        locations = Location.objects.all()
        if not is_superadmin(request.user):
            if not request.user.location_id:
                return Response(
                    {"detail": "Admin user is not assigned to a location."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            locations = locations.filter(pk=request.user.location_id, is_deleted=False)
        elif not include_deleted:
            locations = locations.filter(is_deleted=False)
        serializer = LocationSerializer(locations, many=True)
        return Response(serializer.data)

    def post(self, request):
        if not is_superadmin(request.user):
            return Response(
                {"detail": "Only super admins can create locations."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = LocationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LocationDetailView(AuthenticatedAPIView):
    def get_object(self, pk):
        try:
            return Location.objects.get(pk=pk, is_deleted=False)
        except Location.DoesNotExist:
            return None

    def get(self, request, pk):
        location = self.get_object(pk)
        if not location:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.ADMIN and location.id != request.user.location_id:
            return Response(status=status.HTTP_403_FORBIDDEN)
        serializer = LocationSerializer(location)
        return Response(serializer.data)

    def put(self, request, pk):
        return self.patch(request, pk)

    def patch(self, request, pk):
        if not is_superadmin(request.user):
            return Response(
                {"detail": "Only super admins can update locations."},
                status=status.HTTP_403_FORBIDDEN,
            )
        location = self.get_object(pk)
        if not location:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = LocationSerializer(location, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        if not is_superadmin(request.user):
            return Response(
                {"detail": "Only super admins can delete locations."},
                status=status.HTTP_403_FORBIDDEN,
            )
        location = self.get_object(pk)
        if not location:
            return Response(status=status.HTTP_404_NOT_FOUND)
        location.is_deleted = True
        location.save(update_fields=["is_deleted", "updated_at"])
        return Response(status=status.HTTP_204_NO_CONTENT)


class UserListCreateView(AuthenticatedAPIView):
    def get_queryset(self, request):
        queryset = User.objects.filter(is_deleted=False)
        role = request.query_params.get("role")
        location_id = request.query_params.get("location_id")
        is_active = request.query_params.get("is_active")

        if role:
            queryset = queryset.filter(role=role)
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")

        if request.user.role == User.Role.ADMIN:
            queryset = queryset.filter(location=request.user.location).exclude(
                role=User.Role.SUPERADMIN
            )
        return queryset

    def get(self, request):
        users = self.get_queryset(request)
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data)

    def post(self, request):
        if not is_superadmin(request.user):
            return Response(
                {"detail": "Only super admins can create users."},
                status=status.HTTP_403_FORBIDDEN,
            )
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            output = UserSerializer(user).data
            return Response(output, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserDetailView(AuthenticatedAPIView):
    def get_object(self, pk):
        try:
            return User.objects.get(pk=pk, is_deleted=False)
        except User.DoesNotExist:
            return None

    def get(self, request, pk):
        user = self.get_object(pk)
        if not user:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.ADMIN and user.location != request.user.location:
            return Response(status=status.HTTP_403_FORBIDDEN)
        serializer = UserSerializer(user)
        return Response(serializer.data)

    def put(self, request, pk):
        return self.patch(request, pk)

    def patch(self, request, pk):
        user = self.get_object(pk)
        if not user:
            return Response(status=status.HTTP_404_NOT_FOUND)

        if request.user.role == User.Role.ADMIN and request.user.id != user.id:
            return Response(status=status.HTTP_403_FORBIDDEN)
        if request.user.role == User.Role.ADMIN:
            request.data.setdefault("role", User.Role.ADMIN)

        serializer = UserSerializer(user, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        if request.user.role == User.Role.ADMIN:
            new_location = serializer.validated_data.get("location", user.location)
            if new_location != request.user.location:
                return Response(
                    {"detail": "Admins cannot change their assigned location."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        user = self.get_object(pk)
        if not user:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if not is_superadmin(request.user) and request.user.id != user.id:
            return Response(status=status.HTTP_403_FORBIDDEN)
        user.is_deleted = True
        user.is_active = False
        user.save(update_fields=["is_deleted", "is_active", "updated_at"])
        AuthToken.objects.filter(user=user).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class LoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")
        if not email or not password:
            return Response(
                {"detail": "Email and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(email=email, is_deleted=False)
        except User.DoesNotExist:
            return Response(
                {"detail": "Invalid credentials."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        if not user.check_password(password):
            return Response(
                {"detail": "Invalid credentials."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        if not user.is_active:
            return Response(
                {"detail": "User account is inactive."},
                status=status.HTTP_403_FORBIDDEN,
            )

        AuthToken.objects.filter(user=user).delete()
        token = AuthToken.objects.create(user=user)
        data = UserSerializer(user).data
        data["token"] = token.key
        return Response(data)


class LogoutView(AuthenticatedAPIView):
    def post(self, request):
        AuthToken.objects.filter(user=request.user).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class EmployeeListView(AuthenticatedAPIView):
    def get(self, request):
        queryset = Employee.objects.select_related("location")
        location_id = request.query_params.get("location_id")
        if request.user.role == User.Role.ADMIN:
            queryset = queryset.filter(location=request.user.location)
        elif location_id:
            queryset = queryset.filter(location_id=location_id)
        serializer = EmployeeListSerializer(queryset, many=True)
        return Response(serializer.data)


class EmployeeDetailView(AuthenticatedAPIView):
    def get_object(self, request, pk):
        try:
            employee = Employee.objects.select_related("location").get(pk=pk)
        except Employee.DoesNotExist:
            return None
        if request.user.role == User.Role.ADMIN and employee.location != request.user.location:
            return None
        return employee

    def get(self, request, pk):
        employee = self.get_object(request, pk)
        if not employee:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = EmployeeSerializer(employee)
        return Response(serializer.data)

    def put(self, request, pk):
        return self.patch(request, pk)

    def patch(self, request, pk):
        employee = self.get_object(request, pk)
        if not employee:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = EmployeeUpdateSerializer(employee, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        extra_kwargs = {}
        if request.user.role == User.Role.ADMIN:
            if not request.user.location:
                return Response(
                    {"detail": "Admin user is not assigned to a location."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            new_location = serializer.validated_data.get("location", request.user.location)
            if new_location and new_location != request.user.location:
                return Response(
                    {"detail": "Admins cannot move employees to a different location."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            extra_kwargs["location"] = request.user.location

        employee = serializer.save(**extra_kwargs)

        face_file = request.FILES.get("face_image")
        profile_file = request.FILES.get("profile_photo")
        updated_fields = []

        face_bytes = None
        if face_file:
            face_bytes = face_file.read()
            encoding = get_face_encoding(face_bytes)
            if encoding is None:
                return Response(
                    {"face_image": ["No face detected in provided image."]},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            employee.face_encoding = encoding.tobytes()
            updated_fields.append("face_encoding")

        if profile_file:
            employee.photo = profile_file.read()
            updated_fields.append("photo")
        elif face_bytes and not employee.photo:
            employee.photo = face_bytes
            if "photo" not in updated_fields:
                updated_fields.append("photo")

        if updated_fields:
            employee.save(update_fields=updated_fields)

        return Response(EmployeeSerializer(employee).data)

    def delete(self, request, pk):
        employee = self.get_object(request, pk)
        if not employee:
            return Response(status=status.HTTP_404_NOT_FOUND)
        employee.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class FaceAttendanceView(APIView):
    permission_classes = [permissions.AllowAny]
    def post(self, request):
        serializer = FaceUploadSerializer(data=request.data)
        if not serializer.is_valid():
            logger.warning("Invalid face upload data: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        image = serializer.validated_data["image"]
        uploaded_encoding = get_face_encoding(image.read())

        if uploaded_encoding is None:
            logger.info("No face detected in uploaded image")
            return Response({"error": "No face detected"}, status=status.HTTP_400_BAD_REQUEST)

        employees = Employee.objects.only("id", "name", "face_encoding", "photo")
        user = getattr(request, "user", None)
        if isinstance(user, User) and user.role == User.Role.ADMIN:
            employees = employees.filter(location=user.location)

        known_encodings = []
        employee_map = []

        for emp in employees:
            encoding = np.frombuffer(emp.face_encoding)
            known_encodings.append(encoding)
            employee_map.append(emp)

        matches = face_recognition.compare_faces(
            known_encodings, uploaded_encoding, tolerance=0.45
        )
        if True in matches:
            best_match_index = matches.index(True)
            matched_employee = employee_map[best_match_index]
            #today = timezone.now().date()
            #now = timezone.localtime()
            now = to_ist(timezone.now())
            today = now.date()

            # Create timezone-aware datetime objects for the start and end of today
            today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
            today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))

            logs_today = AttendanceLog.objects.filter(
                employee=matched_employee, timestamp__range=(today_start, today_end)
            )
            has_checkin = any(log.type == "checkin" for log in logs_today)
            has_checkout = any(log.type == "checkout" for log in logs_today)

            if not has_checkin:
                entry_type = "checkin"
                message = f"Welcome, {matched_employee.name.strip()}! Your check-in has been recorded."
            elif not has_checkout:
                entry_type = "checkout"
                message = (
                    f"Good job today, {matched_employee.name.strip()}! Your check-out is complete."
                )
            else:
                logger.info(
                    "Both checkin and checkout already marked for %s",
                    matched_employee.name.strip(),
                )
                return Response(
                    {
                    "status": "Already marked",
                    "message": "You've already completed both check-in and check-out for today.",
                    "employee": matched_employee.name.strip(),
                        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
                    },
                    status=status.HTTP_200_OK,
                )

            AttendanceLog.objects.create(employee=matched_employee, type=entry_type)
            logger.info("%s marked for %s", entry_type.capitalize(), matched_employee.name.strip())

            confidence = round(
                1 - face_recognition.face_distance(
                [known_encodings[best_match_index]], uploaded_encoding
                )[0],
                2,
            )

            photo_base64 = (
                base64.b64encode(matched_employee.photo).decode("utf-8")
                if matched_employee.photo
                else None
            )

            return Response(
                {
                "status": f"{entry_type.capitalize()} successful",
                "message": message,
                "employee": matched_employee.name.strip(),
                "confidence": confidence,
                    "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
                    "photo": f"data:image/jpeg;base64,{photo_base64}"
                    if photo_base64
                    else None,
                },
                status=status.HTTP_200_OK,
            )

        logger.info("Face not recognized")
        return Response({"error": "Face not recognized"}, status=status.HTTP_404_NOT_FOUND)


class RegisterEmployeeView(AuthenticatedAPIView):
    def post(self, request):
        serializer = EmployeeRegisterSerializer(data=request.data, context={"request": request})
        if not serializer.is_valid():
            logger.warning("Invalid registration data: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        name = serializer.validated_data["name"].strip()
        location = serializer.validated_data["location"]
        face_file = serializer.validated_data["face_image"]
        profile_file = serializer.validated_data.get("profile_photo")

        face_bytes = face_file.read()
        encoding = get_face_encoding(face_bytes)

        if encoding is None:
            logger.info("No face detected during registration")
            return Response({"error": "No face detected"}, status=status.HTTP_400_BAD_REQUEST)

        if request.user.role == User.Role.ADMIN:
            if not request.user.location_id:
                return Response(
                    {"error": "Admin user is not assigned to a location."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if location.id != request.user.location_id:
                return Response(
                    {"error": "Admins can only register employees for their own location."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        profile_bytes = profile_file.read() if profile_file else None

        employee = Employee.objects.create(
            name=name,
            location=location,
            face_encoding=encoding.tobytes(),
            photo=profile_bytes or face_bytes,
        )
        logger.info("Employee registered: %s", name)

        return Response(
            {
            "status": "Employee registered",
            "employee_id": employee.id,
                "name": employee.name,
                "location_id": str(location.id),
            },
            status=status.HTTP_201_CREATED,
        )


class AttendanceSummaryView(AuthenticatedAPIView):
    def get(self, request):
        # Accept optional start_date and end_date query params (YYYY-MM-DD).
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")

        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except Exception:
                return Response({"error": "Invalid date format, expected YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)
        elif start_date_str and not end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = start_date
            except Exception:
                return Response({"error": "Invalid date format, expected YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Use timezone-aware current date
            current_datetime = timezone.now()
            start_date = end_date = current_datetime.date()

        if start_date > end_date:
            return Response({"error": "start_date cannot be after end_date"}, status=status.HTTP_400_BAD_REQUEST)

        employees = Employee.objects.prefetch_related("attendancelog_set")
        if request.user.role == User.Role.ADMIN:
            employees = employees.filter(location=request.user.location)

        summary = []
        # iterate dates in range and collect per-employee rows per date
        delta_days = (end_date - start_date).days
        for single_day_offset in range(delta_days + 1):
            current_date = start_date + timedelta(days=single_day_offset)
            # Create timezone-aware datetime objects for the start and end of the day
            day_start = timezone.make_aware(datetime.combine(current_date, datetime.min.time()))
            day_end = timezone.make_aware(datetime.combine(current_date, datetime.max.time()))

            for emp in employees:
                logs = emp.attendancelog_set.filter(timestamp__range=(day_start, day_end))
                checkin_time = logs.filter(type="checkin").aggregate(Min("timestamp"))["timestamp__min"]
                checkout_time = logs.filter(type="checkout").aggregate(Max("timestamp"))["timestamp__max"]

                if checkin_time and checkout_time:
                    duration_seconds = (checkout_time - checkin_time).total_seconds()
                    hours = int(duration_seconds // 3600)
                    minutes = int((duration_seconds % 3600) // 60)
                    duration_str = f"{hours:02d}:{minutes:02d}"
                else:
                    duration_str = None

                summary.append(
                    {
                        "employee": emp.name,
                        "date": current_date.strftime("%Y-%m-%d"),
                        "checkin": to_ist(checkin_time).strftime("%H:%M:%S") if checkin_time else None,
                        "checkout": to_ist(checkout_time).strftime("%H:%M:%S") if checkout_time else None,
                        "duration": duration_str,
                    }
                )

        return Response(summary)


class AttendanceSummaryExportView(AuthenticatedAPIView):
    def get(self, request):
        # Support optional start_date and end_date params (YYYY-MM-DD)
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")

        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except Exception:
                return Response({"error": "Invalid date format, expected YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)
        elif start_date_str and not end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = start_date
            except Exception:
                return Response({"error": "Invalid date format, expected YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Use timezone-aware current date
            current_datetime = timezone.now()
            start_date = end_date = current_datetime.date()

        if start_date > end_date:
            return Response({"error": "start_date cannot be after end_date"}, status=status.HTTP_400_BAD_REQUEST)

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"

        ws.append(["Employee", "Date", "Check-in", "Check-out", "Duration"])

        employees = Employee.objects.prefetch_related("attendancelog_set")
        if request.user.role == User.Role.ADMIN:
            employees = employees.filter(location=request.user.location)

        delta_days = (end_date - start_date).days
        for single_day_offset in range(delta_days + 1):
            current_date = start_date + timedelta(days=single_day_offset)
            # Create timezone-aware datetime objects for the start and end of the day
            day_start = timezone.make_aware(datetime.combine(current_date, datetime.min.time()))
            day_end = timezone.make_aware(datetime.combine(current_date, datetime.max.time()))

            for emp in employees:
                logs = emp.attendancelog_set.filter(timestamp__range=(day_start, day_end))
                checkin_time = logs.filter(type="checkin").aggregate(Min("timestamp"))["timestamp__min"]
                checkout_time = logs.filter(type="checkout").aggregate(Max("timestamp"))["timestamp__max"]

                if checkin_time and checkout_time:
                    duration_seconds = (checkout_time - checkin_time).total_seconds()
                    hours = int(duration_seconds // 3600)
                    minutes = int((duration_seconds % 3600) // 60)
                    duration_str = f"{hours:02d}:{minutes:02d}"
                else:
                    duration_str = ""

                ws.append([
                    emp.name,
                    current_date.strftime("%Y-%m-%d"),
                    to_ist(checkin_time).strftime("%H:%M:%S") if checkin_time else "",
                    to_ist(checkout_time).strftime("%H:%M:%S") if checkout_time else "",
                    duration_str,
                ])

        filename = f"attendance_summary_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
        wb.save(filepath)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + filename)
        return Response({"file_url": file_url})


class MonthlyAttendanceStatusView(AuthenticatedAPIView):
    def get(self, request):
        month = request.query_params.get("month")
        if not month:
            return Response(
                {"error": "Month is required in YYYY-MM format"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            year, month_num = map(int, month.split("-"))
            start_date = datetime(year, month_num, 1).date()
            end_date = datetime(year, month_num, monthrange(year, month_num)[1]).date()
        except Exception:
            return Response(
                {"error": "Invalid month format"}, status=status.HTTP_400_BAD_REQUEST
            )

        date_range = [
            start_date + timedelta(days=i)
            for i in range((end_date - start_date).days + 1)
        ]
        employees = Employee.objects.select_related("location").all()
        if request.user.role == User.Role.ADMIN:
            employees = employees.filter(location=request.user.location)

        # logs = AttendanceLog.objects.filter(
        #     timestamp__date__range=(start_date, end_date)
        # )
        day_start = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
        day_end = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))

        logs = AttendanceLog.objects.filter(timestamp__range=(day_start, day_end))

        print(logs)
        if request.user.role == User.Role.ADMIN:
            logs = logs.filter(employee__location=request.user.location)
            
        attendance_map = defaultdict(lambda: defaultdict(lambda: "-"))
        for log in logs:
            local_date = to_ist(log.timestamp).date()
            attendance_map[log.employee_id][local_date] = "P"

        summary = []
        for emp in employees:
            row = {"name": emp.name}
            for day in date_range:
                status_code = attendance_map[emp.id].get(
                    day, "A" if emp.id in attendance_map else "-"
                )
                row[day.strftime("%d-%b")] = status_code
            summary.append(row)

        return Response(summary)


class MonthlyAttendanceStatusExportView(AuthenticatedAPIView):
    def get(self, request):
        month = request.query_params.get("month")
        if not month:
            return Response({"error": "Month is required in YYYY-MM format"}, status=400)

        try:
            year, month_num = map(int, month.split("-"))
            start_date = datetime(year, month_num, 1).date()
            end_date = datetime(year, month_num, monthrange(year, month_num)[1]).date()
        except Exception:
            return Response({"error": "Invalid month format"}, status=400)

        date_range = [
            start_date + timedelta(days=i)
            for i in range((end_date - start_date).days + 1)
        ]
        employees = Employee.objects.select_related("location").all()
        if request.user.role == User.Role.ADMIN:
            employees = employees.filter(location=request.user.location)

        # logs = AttendanceLog.objects.filter(
        #     timestamp__date__range=(start_date, end_date)
        # )
        
        day_start = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
        day_end = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))

        logs = AttendanceLog.objects.filter(timestamp__range=(day_start, day_end))

        if request.user.role == User.Role.ADMIN:
            logs = logs.filter(employee__location=request.user.location)

        attendance_map = {}
        for log in logs:
            local_date = to_ist(log.timestamp).date()
            key = (log.employee_id, local_date)
            attendance_map[key] = "P"


        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {month}"

        header = ["Name"] + [d.strftime("%d-%b") for d in date_range] + [
            "Present",
            "Absent",
        ]
        ws.append(header)

        for emp in employees:
            row = [emp.name]
            present_count = 0
            absent_count = 0

            for d in date_range:
                key = (emp.id, d)
                status_code = attendance_map.get(
                    key,
                    "A" if any(k[0] == emp.id for k in attendance_map) else "-",
                )
                row.append(status_code)
                if status_code == "P":
                    present_count += 1
                elif status_code == "A":
                    absent_count += 1

            row.append(present_count)
            row.append(absent_count)
            ws.append(row)

        filename = f"monthly_attendance_{month}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)
        # Ensure media directory exists before saving the workbook
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
        wb.save(filepath)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + filename)
        return Response({"file_url": file_url})


class GeneratePayrollView(AuthenticatedAPIView):
    def post(self, request):
        month = request.data.get("month")
        if not month:
            return Response({"error": "Month is required"}, status=400)

        year, month_num = map(int, month.split("-"))
        start_date = datetime(year, month_num, 1).date()
        end_date = datetime(year, month_num, monthrange(year, month_num)[1]).date()

        employees = Employee.objects.all()
        if request.user.role == User.Role.ADMIN:
            employees = employees.filter(location=request.user.location)

        logs = AttendanceLog.objects.filter(timestamp__date__range=(start_date, end_date))
        if request.user.role == User.Role.ADMIN:
            logs = logs.filter(employee__location=request.user.location)

        attendance_map = {}
        for log in logs:
            key = (log.employee_id, to_ist(log.timestamp).date())
            attendance_map[key] = "P"

        for emp in employees:
            present = sum(
                1
                for day in range((end_date - start_date).days + 1)
                if attendance_map.get((emp.id, start_date + timedelta(days=day))) == "P"
            )
            absent = sum(
                1
                for day in range((end_date - start_date).days + 1)
                if attendance_map.get((emp.id, start_date + timedelta(days=day))) == "A"
            )

            base_salary = emp.base_salary or 0
            deduction_per_day = emp.deduction_per_day or 0
            deductions = absent * deduction_per_day
            pf_deduction = (base_salary * Decimal("0.12")).quantize(Decimal("0.01"))
            esi_deduction = (base_salary * Decimal("0.0175")).quantize(
                Decimal("0.01")
            )
            net_pay = base_salary - deductions - pf_deduction - esi_deduction

            PayrollRecord.objects.create(
                employee=emp,
                month=month,
                present_days=present,
                absent_days=absent,
                base_salary=base_salary,
                deduction_per_day=deduction_per_day,
                deductions=deductions,
                pf_deduction=pf_deduction,
                esi_deduction=esi_deduction,
                net_pay=net_pay,
            )

        return Response({"status": f"Payroll generated for {month}"})


class PayrollExportView(AuthenticatedAPIView):
    def get(self, request):
        month = request.query_params.get("month")
        if not month:
            return Response({"error": "Month is required in YYYY-MM format"}, status=400)

        records = PayrollRecord.objects.filter(month=month)
        if request.user.role == User.Role.ADMIN:
            records = records.filter(employee__location=request.user.location)

        if not records.exists():
            return Response(
                {"error": "No payroll records found for this month"}, status=404
            )

        wb = Workbook()
        ws = wb.active
        ws.title = f"Payroll {month}"

        ws.append(
            [
                "Employee",
                "Present Days",
                "Absent Days",
                "Base Salary",
                "Deduction/Day",
                "Deductions",
                "PF",
                "ESI",
                "Net Pay",
            ]
        )

        for record in records:
            ws.append(
                [
                record.employee.name,
                record.present_days,
                record.absent_days,
                float(record.base_salary),
                float(record.deduction_per_day),
                float(record.deductions),
                float(record.pf_deduction),
                float(record.esi_deduction),
                float(record.net_pay),
                ]
            )

        filename = f"payroll_{month}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)
        # Ensure media directory exists before saving the workbook
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
        wb.save(filepath)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + filename)
        return Response({"file_url": file_url})